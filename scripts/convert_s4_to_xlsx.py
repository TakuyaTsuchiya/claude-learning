#!/usr/bin/env python3
"""Convert S4-V1 ~ S4-V5 JSON files to CC_セクション4.xlsx"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "/Users/tchytky/Desktop/claude-learning/gems-outputs"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "CC_セクション4.xlsx")

# Color palette
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")  # medium blue
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")  # very light blue
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(name="Yu Gothic", bold=True, color="1F4E79", size=10)
BODY_FONT = Font(name="Yu Gothic", size=10)
TITLE_FONT = Font(name="Yu Gothic", bold=True, color="1F4E79", size=14)

thin = Side(style="thin", color="BBBBBB")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_header(cell, text, font=None, fill=None):
    cell.value = text
    cell.font = font or HEADER_FONT
    cell.fill = fill or HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def apply_body(cell, text, fill=None, bold=False):
    cell.value = text
    cell.font = Font(name="Yu Gothic", bold=bold, size=10)
    cell.fill = fill or WHITE_FILL
    cell.alignment = WRAP
    cell.border = THIN_BORDER


def write_metadata_sheet(wb, data, video_num):
    ws = wb.create_sheet(title=f"V{video_num}_メタデータ")
    meta = data["metadata"]

    # Title row
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"【セクション4 動画{video_num}】メタデータ"
    c.font = TITLE_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    fields = [
        ("モジュール", meta.get("module", "")),
        ("セクション", meta.get("section", "")),
        ("セクションタイトル", meta.get("section_title", "")),
        ("動画タイトル", meta.get("title", "")),
        ("尺", meta.get("duration", "")),
        ("モード", meta.get("mode", "")),
        ("学習ゴール", meta.get("goal", "")),
        ("最重要トピック", meta.get("topics", "")),
        ("つまずき予測", meta.get("stumbling", "")),
        ("画面収録方針", meta.get("screen_policy", "")),
        ("ポジショニング", meta.get("positioning", "")),
        ("参照情報", meta.get("references", "")),
    ]

    for i, (label, value) in enumerate(fields, start=2):
        ws.merge_cells(f"C{i}:D{i}")
        apply_body(ws[f"A{i}"], "", fill=SECTION_FILL)
        apply_body(ws[f"B{i}"], label, bold=True, fill=SECTION_FILL)
        apply_body(ws[f"C{i}"], value)
        ws[f"B{i}"].font = LABEL_FONT

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20

    for i in range(2, len(fields) + 2):
        ws.row_dimensions[i].height = 40


def write_materials_sheet(wb, data, video_num):
    ws = wb.create_sheet(title=f"V{video_num}_準備物")
    materials = data["materials"]

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"【セクション4 動画{video_num}】準備物リスト"
    c.font = TITLE_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    headers = ["種別", "名称", "用途", "挿入タイミング", "ファイル名例"]
    for col, h in enumerate(headers, start=1):
        apply_header(ws.cell(row=2, column=col), h)

    for i, mat in enumerate(materials):
        row = i + 3
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        apply_body(ws.cell(row=row, column=1), mat.get("type", ""), fill=fill)
        apply_body(ws.cell(row=row, column=2), mat.get("name", ""), fill=fill)
        apply_body(ws.cell(row=row, column=3), mat.get("purpose", ""), fill=fill)
        apply_body(ws.cell(row=row, column=4), mat.get("timing", ""), fill=fill)
        apply_body(ws.cell(row=row, column=5), mat.get("filename", ""), fill=fill)
        ws.row_dimensions[row].height = 35

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30


def write_script_sheet(wb, data, video_num):
    ws = wb.create_sheet(title=f"V{video_num}_台本")
    script = data["script"]
    title = data["metadata"].get("title", "")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"【セクション4 動画{video_num}】台本　{title}"
    c.font = TITLE_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    headers = ["開始", "終了", "トピック", "演出指示", "台本（セリフ）", "参考資料"]
    for col, h in enumerate(headers, start=1):
        apply_header(ws.cell(row=2, column=col), h)

    for i, line in enumerate(script):
        row = i + 3
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        apply_body(ws.cell(row=row, column=1), line.get("start", ""), fill=fill)
        apply_body(ws.cell(row=row, column=2), line.get("end", ""), fill=fill)
        apply_body(ws.cell(row=row, column=3), line.get("topic", ""), fill=fill, bold=True)
        apply_body(ws.cell(row=row, column=4), line.get("direction", ""), fill=fill)
        apply_body(ws.cell(row=row, column=5), line.get("content", ""), fill=fill)
        apply_body(ws.cell(row=row, column=6), line.get("reference", ""), fill=fill)
        ws.row_dimensions[row].height = 120

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 28


def write_actions_sheet(wb, data, video_num):
    ws = wb.create_sheet(title=f"V{video_num}_アクション")
    actions = data["actions"]

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"【セクション4 動画{video_num}】講師アクション一覧"
    c.font = TITLE_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    headers = ["アクション内容", "完了", "メモ"]
    for col, h in enumerate(headers, start=1):
        apply_header(ws.cell(row=2, column=col), h)

    for i, action in enumerate(actions):
        row = i + 3
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        apply_body(ws.cell(row=row, column=1), action.get("action", ""), fill=fill)
        apply_body(ws.cell(row=row, column=2), action.get("done", "□"), fill=fill)
        apply_body(ws.cell(row=row, column=3), action.get("memo", ""), fill=fill)
        ws.row_dimensions[row].height = 45

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for video_num in range(1, 6):
        json_path = os.path.join(OUTPUT_DIR, f"S4-V{video_num}.json")
        print(f"Loading {json_path}...")
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        write_metadata_sheet(wb, data, video_num)
        write_materials_sheet(wb, data, video_num)
        write_script_sheet(wb, data, video_num)
        write_actions_sheet(wb, data, video_num)
        print(f"  V{video_num} sheets written.")

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
