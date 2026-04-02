import json
import openpyxl
from openpyxl.styles import Font, Alignment

OUTPUT_DIR = "/Users/tchytky/Desktop/claude-learning/gems-outputs"

def set_header(cell, value):
    cell.value = value
    cell.font = Font(bold=True)

def set_wrap(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def write_video_sheet(ws, data):
    meta = data["metadata"]
    materials = data["materials"]
    script = data["script"]
    actions = data["actions"]

    # --- Column widths (same as S1) ---
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 15

    # --- Metadata section ---
    meta_rows = [
        ("項目", "内容"),
        ("モジュール", meta.get("module", "")),
        ("セクション", meta.get("section", "")),
        ("動画タイトル", meta.get("title", "")),
        ("尺", meta.get("duration", "")),
        ("モード", meta.get("mode", "")),
        ("ゴール", meta.get("goal", "")),
        ("重要トピック", meta.get("topics", "")),
        ("受講者のつまずき", meta.get("stumbling", "")),
        ("画面操作方針", meta.get("screen_policy", "")),
        ("位置付け方針", meta.get("positioning", "")),
        ("参照一覧", meta.get("references", "")),
    ]
    for r_idx, (k, v) in enumerate(meta_rows, 1):
        cell_a = ws.cell(row=r_idx, column=1, value=k)
        cell_b = ws.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
        set_wrap(cell_b)

    # --- Materials section ---
    mat_start = len(meta_rows) + 1  # row 13
    headers = ("種別", "名称", "用途", "挿入タイミング", "ファイル名例")
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=mat_start, column=col_idx, value=h)
        c.font = Font(bold=True)
    for i, m in enumerate(materials):
        r = mat_start + 1 + i
        ws.cell(row=r, column=1, value=m.get("type", ""))
        ws.cell(row=r, column=2, value=m.get("name", ""))
        cell_c = ws.cell(row=r, column=3, value=m.get("purpose", ""))
        set_wrap(cell_c)
        ws.cell(row=r, column=4, value=m.get("timing", ""))
        ws.cell(row=r, column=5, value=m.get("filename", ""))

    # blank row before script
    script_header_row = mat_start + 1 + len(materials) + 1  # +1 blank

    # --- Script section ---
    sc_headers = ("開始", "終了", "話題", "シーン演出", "話す内容", "引用")
    for col_idx, h in enumerate(sc_headers, 1):
        c = ws.cell(row=script_header_row, column=col_idx, value=h)
        c.font = Font(bold=True)
    for i, s in enumerate(script):
        r = script_header_row + 1 + i
        ws.cell(row=r, column=1, value=s.get("start", ""))
        ws.cell(row=r, column=2, value=s.get("end", ""))
        ws.cell(row=r, column=3, value=s.get("topic", ""))
        cell_d = ws.cell(row=r, column=4, value=s.get("direction", ""))
        set_wrap(cell_d)
        cell_e = ws.cell(row=r, column=5, value=s.get("content", ""))
        set_wrap(cell_e)
        ws.cell(row=r, column=6, value=s.get("reference", ""))

    # blank row before actions
    action_header_row = script_header_row + 1 + len(script) + 1

    # --- Actions section ---
    ac_headers = ("アクション", "完了", "メモ")
    for col_idx, h in enumerate(ac_headers, 1):
        c = ws.cell(row=action_header_row, column=col_idx, value=h)
        c.font = Font(bold=True)
    for i, a in enumerate(actions):
        r = action_header_row + 1 + i
        cell_a = ws.cell(row=r, column=1, value=a.get("action", ""))
        set_wrap(cell_a)
        ws.cell(row=r, column=2, value=a.get("done", "□"))
        cell_c = ws.cell(row=r, column=3, value=a.get("memo", ""))
        set_wrap(cell_c)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for v in range(1, 5):
        json_path = f"{OUTPUT_DIR}/S3-V{v}.json"
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        ws = wb.create_sheet(title=f"動画{v}")
        write_video_sheet(ws, data)

    out_path = f"{OUTPUT_DIR}/CC_セクション3.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
